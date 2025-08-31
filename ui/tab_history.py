# -*- coding: utf-8 -*-
from datetime import datetime
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QDateEdit, QLineEdit, QComboBox,
    QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog
)

import pandas as pd

# Estes imports são opcionais: se openpyxl faltar, a exportação continua sem formatação
try:
    from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

from models import pretty_money, Trade, Wallet


class TabHistory(QWidget):
    """
    Histórico de trades (da carteira selecionada), com filtros,
    apagar trade, e EXPORTAÇÃO para Excel (Trades + Estatísticas).
    """
    COLS = [
        ("id", "ID"),
        ("created_at", "Data"),
        ("wallet", "Carteira"),
        ("symbol", "Paridade"),
        ("direction", "Direção"),
        ("entry_price", "Entrada"),
        ("stop_loss", "SL"),
        ("take_profit", "TP"),
        ("position_size", "Quantidade"),
        ("position_value", "Valor posição"),
        ("risk_amount", "Risco $"),
        ("risk_pct_of_balance", "Risco % saldo"),
        ("status", "Estado"),
        ("exit_price", "Saída"),
        ("pnl_abs", "PnL $"),
        ("pnl_pct", "PnL %"),
        ("result", "Resultado"),
        ("close_reason", "Fechado Como"),
        ("reason", "Razão"),
    ]

    def __init__(self, app):
        super().__init__()
        self.app = app
        self._rows_cache = []  # cache dos trades filtrados exibidos
        self.build()

    def build(self):
        self.setStyleSheet("""
            QLabel { font-size: 12pt; }
            QComboBox, QDateEdit, QLineEdit, QPushButton, QTableWidget { font-size: 11pt; }
        """)
        v = QVBoxLayout(self)

        # ---- Filtros topo
        filt = QHBoxLayout()
        filt.addWidget(QLabel("De:"))
        self.dt_from = QDateEdit()
        self.dt_from.setCalendarPopup(True)
        self.dt_from.setDate(QDate(2000, 1, 1))
        filt.addWidget(self.dt_from)

        filt.addWidget(QLabel("Até:"))
        self.dt_to = QDateEdit()
        self.dt_to.setCalendarPopup(True)
        self.dt_to.setDate(QDate.currentDate())
        filt.addWidget(self.dt_to)

        filt.addWidget(QLabel("Paridade:"))
        self.ed_f_symbol = QLineEdit()
        self.ed_f_symbol.setPlaceholderText("ex.: BTCUSDT")
        filt.addWidget(self.ed_f_symbol)

        filt.addWidget(QLabel("Estado:"))
        self.cmb_status = QComboBox()
        self.cmb_status.addItems(["Todos", "Open", "Closed"])
        filt.addWidget(self.cmb_status)

        self.btn_apply = QPushButton("Aplicar")
        self.btn_apply.clicked.connect(self.refresh_table)
        filt.addWidget(self.btn_apply)

        filt.addStretch()

        self.btn_export = QPushButton("Exportar Excel")
        self.btn_export.setStyleSheet("font-weight:600;")
        self.btn_export.clicked.connect(self.export_to_excel)
        filt.addWidget(self.btn_export)

        self.btn_delete = QPushButton("Apagar Trade")
        self.btn_delete.clicked.connect(self.delete_selected_trade)
        filt.addWidget(self.btn_delete)

        v.addLayout(filt)

        # ---- Tabela
        self.tbl = QTableWidget(0, len(self.COLS))
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setSelectionMode(QTableWidget.SingleSelection)
        self.tbl.setSortingEnabled(True)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setAlternatingRowColors(True)
        for i, (_, header) in enumerate(self.COLS):
            self.tbl.setHorizontalHeaderItem(i, QTableWidgetItem(header))
        v.addWidget(self.tbl)

        # primeira carga
        self.refresh_table()

        # sinais para refazer ao mudar filtros
        self.dt_from.dateChanged.connect(self.refresh_table)
        self.dt_to.dateChanged.connect(self.refresh_table)
        self.ed_f_symbol.textChanged.connect(self.refresh_table)
        self.cmb_status.currentIndexChanged.connect(self.refresh_table)

    # ------------------------------------------------------------------
    # Helpers de dados/filters
    # ------------------------------------------------------------------
    def _date_ok(self, iso: str) -> bool:
        """Verifica se a data ISO está dentro do intervalo (De/Até)."""
        try:
            d = datetime.fromisoformat(iso)
        except Exception:
            return True  # sem data bem formatada, não filtra
        dmin = datetime(self.dt_from.date().year(), self.dt_from.date().month(), self.dt_from.date().day())
        dmax = datetime(self.dt_to.date().year(), self.dt_to.date().month(), self.dt_to.date().day(), 23, 59, 59)
        return dmin <= d <= dmax

    def _wallet_name_by_id(self, wid: str) -> str:
        """Resolve nome da carteira localmente, sem depender de métodos no app."""
        try:
            w = self.app.ds.wallets.get(wid)
            return w.name if w else ""
        except Exception:
            return ""

    def _filter_trades(self):
        """Aplica filtros aos trades da carteira selecionada."""
        w: Wallet = self.app.current_wallet()
        if not w:
            return []

        symbol_f = self.ed_f_symbol.text().strip().upper()
        status_f = self.cmb_status.currentText()
        rows = []
        for t in self.app.ds.trades.values():
            if t.wallet_id != w.id:
                continue
            if symbol_f and symbol_f not in (t.symbol or "").upper():
                continue
            if t.created_at and not self._date_ok(t.created_at):
                continue
            if status_f != "Todos" and (t.status or "Open") != status_f:
                continue
            rows.append(t)
        rows.sort(key=lambda x: (x.created_at or ""))
        return rows

    def refresh_table(self):
        """Repovoa a tabela com base nos filtros atuais."""
        rows = self._filter_trades()
        self._rows_cache = rows[:]  # guarda para o export

        self.tbl.setSortingEnabled(False)
        self.tbl.setRowCount(0)

        for t in rows:
            r = self.tbl.rowCount()
            self.tbl.insertRow(r)
            vals = self._row_values(t)
            for c, val in enumerate(vals):
                it = QTableWidgetItem(val)
                if c in (5, 6, 7, 8, 9, 10, 11, 13, 14, 15):
                    it.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                self.tbl.setItem(r, c, it)

        self.tbl.resizeColumnsToContents()
        self.tbl.setSortingEnabled(True)

    def _row_values(self, t: Trade):
        """Converte o trade em uma linha de texto para a tabela."""
        wallet_name = self._wallet_name_by_id(t.wallet_id)
        def f2(x): return "" if x is None else f"{float(x):.2f}"
        def f4(x): return "" if x is None else f"{float(x):.4f}"
        return [
            t.id or "",
            (t.created_at or "").replace("T", " "),
            wallet_name or "",
            t.symbol or "",
            t.direction or "",
            f2(t.entry_price),
            f2(t.stop_loss),
            f2(t.take_profit),
            f4(t.position_size),
            f2(t.position_value),
            f2(t.risk_amount),
            f2(t.risk_pct_of_balance),
            t.status or "",
            f2(t.exit_price) if t.exit_price is not None else "",
            f2(t.pnl_abs) if t.pnl_abs is not None else "",
            f2(t.pnl_pct) if t.pnl_pct is not None else "",
            t.result or "",
            t.close_reason or "",
            t.reason or "",
        ]

    def get_selected_trade(self):
        row = self.tbl.currentRow()
        if row < 0:
            return None
        tid = self.tbl.item(row, 0).text()
        return self.app.ds.trades.get(tid)

    def delete_selected_trade(self):
        t = self.get_selected_trade()
        if not t:
            QMessageBox.warning(self, "Apagar", "Seleciona um trade.")
            return
        if QMessageBox.question(self, "Confirmar", f"Apagar trade {t.id} ({t.symbol})?",
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        self.app.ds.delete_trade(t.id)
        self.refresh_table()
        self.app.refresh_all()

    # ------------------------------------------------------------------
    # Exportação para Excel
    # ------------------------------------------------------------------
    def export_to_excel(self):
        """Exporta os trades atualmente listados + estatísticas para um .xlsx."""
        import traceback
        try:
            if not self._rows_cache:
                QMessageBox.information(self, "Exportar", "Não há dados para exportar.")
                return

            # diálogo guardar
            path, _ = QFileDialog.getSaveFileName(self, "Guardar Excel",
                                                  "Tradeiros_Historico.xlsx",
                                                  "Excel (*.xlsx)")
            if not path:
                return

            # DataFrame de Trades (carteira atual)
            trades = self._rows_cache
            w: Wallet = self.app.current_wallet()
            wallet_name = w.name if w else ""

            def as_dict(t: Trade):
                def as_float(v, default=0.0):
                    try:
                        return float(v)
                    except Exception:
                        return default
                return dict(
                    ID=t.id,
                    Data=(t.created_at or "").replace("T", " "),
                    Carteira=self._wallet_name_by_id(t.wallet_id),
                    Paridade=t.symbol,
                    Direção=t.direction,
                    Entrada=as_float(t.entry_price, 0.0),
                    SL=as_float(t.stop_loss, 0.0),
                    TP=as_float(t.take_profit, 0.0),
                    Quantidade=as_float(t.position_size, 0.0),
                    ValorPos=as_float(t.position_value, 0.0),
                    RiscoUSD=as_float(t.risk_amount, 0.0),
                    RiscoPct=as_float(t.risk_pct_of_balance, 0.0),
                    Estado=t.status,
                    Saída=(None if t.exit_price is None else as_float(t.exit_price, 0.0)),
                    PnL=(None if t.pnl_abs is None else as_float(t.pnl_abs, 0.0)),
                    PnLPct=(None if t.pnl_pct is None else as_float(t.pnl_pct, 0.0)),
                    Resultado=t.result,
                    FechadoComo=t.close_reason,
                    Razão=t.reason,
                )

            df = pd.DataFrame([as_dict(t) for t in trades])

            # KPIs carteira selecionada e globais
            stats_wallet = self._compute_stats_for_wallet(w)
            stats_global = self._compute_stats_global()

            # escrever
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                # Trades
                df.to_excel(writer, index=False, sheet_name="Trades")
                if _HAS_OPENPYXL:
                    self._style_trades_sheet(writer.book["Trades"], df)

                # Estatísticas
                ws_stats = writer.book.create_sheet("Estatísticas")
                ws_stats["A1"] = "Carteira selecionada"
                if _HAS_OPENPYXL:
                    ws_stats["A1"].font = Font(bold=True, size=12)
                self._write_stats_block(ws_stats, 2, 1, stats_wallet)

                ws_stats["A10"] = "Global (todas as carteiras)"
                if _HAS_OPENPYXL:
                    ws_stats["A10"].font = Font(bold=True, size=12)
                self._write_stats_block(ws_stats, 11, 1, stats_global)

            QMessageBox.information(self, "Exportar Excel", f"Ficheiro guardado:\n{path}")

        except Exception as e:
            msg = "".join(traceback.format_exc())
            QMessageBox.critical(self, "Erro ao exportar", f"Ocorreu um erro:\n{e}\n\nDetalhes:\n{msg}")

    # ---- helpers de estatística para export ----
    def _compute_stats_for_wallet(self, w: Wallet):
        if not w:
            return {}
        ts = [t for t in self.app.ds.trades.values() if t.wallet_id == w.id]
        return self._compute_stats(ts, w.initial_balance or 0.0)

    def _compute_stats_global(self):
        wallets = list(self.app.ds.wallets.values())
        initial_total = sum((w.initial_balance or 0.0) for w in wallets) if wallets else 0.0
        ts = list(self.app.ds.trades.values())
        return self._compute_stats(ts, initial_total)

    def _compute_stats(self, trades, initial_balance: float):
        closed = [t for t in trades if (t.status or "Open") == "Closed"]
        winners = [t for t in closed if (t.pnl_abs or 0) > 0]
        losers = [t for t in closed if (t.pnl_abs or 0) < 0]
        be = [t for t in closed if (t.pnl_abs or 0) == 0]
        pnl_total = sum((t.pnl_abs or 0.0) for t in closed)
        total = len(trades)
        winrate = (len(winners) / len(closed) * 100.0) if closed else 0.0
        current_balance = (initial_balance or 0.0) + pnl_total
        growth_pct = ((current_balance / initial_balance - 1) * 100.0) if initial_balance and initial_balance > 0 else 0.0
        return dict(
            total_trades=total,
            closed_trades=len(closed),
            open_trades=total - len(closed),
            winners=len(winners),
            losers=len(losers),
            breakeven=len(be),
            winrate_pct=winrate,
            pnl_total=pnl_total,
            initial_balance=initial_balance or 0.0,
            current_balance=current_balance,
            growth_pct=growth_pct
        )

    def _write_stats_block(self, ws, start_row: int, start_col: int, s: dict):
        labels = [
            ("Total de trades", "total_trades"),
            ("Fechados", "closed_trades"),
            ("Abertos", "open_trades"),
            ("Winners", "winners"),
            ("Losers", "losers"),
            ("Break-even", "breakeven"),
            ("Winrate %", "winrate_pct"),
            ("PnL total ($)", "pnl_total"),
            ("Saldo inicial ($)", "initial_balance"),
            ("Saldo atual ($)", "current_balance"),
            ("Crescimento %", "growth_pct"),
        ]
        r = start_row
        money_keys = {"pnl_total", "initial_balance", "current_balance"}
        pct_keys = {"winrate_pct", "growth_pct"}

        for label, key in labels:
            ws.cell(row=r, column=start_col, value=label)
            # deixar negrito se openpyxl disponível
            if _HAS_OPENPYXL:
                ws.cell(row=r, column=start_col).font = Font(bold=True)

            val = s.get(key, 0.0)
            cell = ws.cell(row=r, column=start_col + 1, value=val)

            if _HAS_OPENPYXL:
                if key in money_keys:
                    cell.number_format = u'"$"#,##0.00'
                elif key in pct_keys:
                    try:
                        cell.value = float(val) / 100.0
                    except Exception:
                        pass
                    cell.number_format = "0.00%"
            r += 1

        if _HAS_OPENPYXL:
            try:
                from openpyxl.utils import get_column_letter
                for c in range(start_col, start_col + 2):
                    ws.column_dimensions[get_column_letter(c)].width = 24
            except Exception:
                pass

    # ---- estilo da folha de Trades (se openpyxl presente) ----
    def _style_trades_sheet(self, ws, df):
        if not _HAS_OPENPYXL:
            return
        try:
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # cinza escuro
            header_font = Font(color="FFFFFF", bold=True)
            for c in range(1, df.shape[1] + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center")
                # larguras
                if c in (1, 3, 4, 5, 18):
                    width = 14
                elif c in (2,):
                    width = 20
                else:
                    width = 13
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(c)].width = width

            # filtros + freeze header
            from openpyxl.utils import get_column_letter
            ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0]+1}"
            ws.freeze_panes = "A2"

            # estilos nomeados (registar só se necessário)
            def register_style(style):
                try:
                    ws.parent.add_named_style(style)
                except ValueError:
                    pass

            money = NamedStyle(name="Money_Style")
            money.number_format = u'"$"#,##0.00'
            pct = NamedStyle(name="Pct_Style")
            pct.number_format = "0.00%"

            register_style(money)
            register_style(pct)

            # mapear colunas
            col_map = {name: idx+1 for idx, name in enumerate(df.columns)}

            # aplicar formatos
            for r in range(2, df.shape[0] + 2):
                for nm in ("Entrada", "SL", "TP", "ValorPos", "RiscoUSD", "Saída", "PnL"):
                    c = col_map.get(nm)
                    if c:
                        ws.cell(row=r, column=c).style = "Money_Style"
                for nm in ("RiscoPct", "PnLPct"):
                    c = col_map.get(nm)
                    if c:
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, (int, float)) and v > 1.0:
                            ws.cell(row=r, column=c).value = v / 100.0
                        ws.cell(row=r, column=c).style = "Pct_Style"

            # gradiente no PnL
            if df.shape[0] > 0 and "PnL" in df.columns:
                c = col_map["PnL"]
                rng = f"{get_column_letter(c)}2:{get_column_letter(c)}{df.shape[0]+1}"
                safe_min = float(pd.to_numeric(df['PnL'], errors='coerce').fillna(0).min())
                safe_max = float(pd.to_numeric(df['PnL'], errors='coerce').fillna(0).max())
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type='num', start_value=min(safe_min, 0),
                        mid_type='num', mid_value=0,
                        end_type='num', end_value=max(safe_max, 0),
                        start_color='FCA5A5', mid_color='FFFFFF', end_color='86EFAC'
                    )
                )
        except Exception:
            # qualquer erro de styling é ignorado; o ficheiro continua válido
            pass
